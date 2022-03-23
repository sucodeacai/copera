from colorama import Back, Fore, init
import openpyxl 
from openpyxl.styles import Font, PatternFill, Border
from openpyxl import formatting
from openpyxl.formatting.rule import Rule
from openpyxl import styles

from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import  FormulaRule
from copy import copy



 

wb = openpyxl.load_workbook(filename="files/copera.xlsx")

#Show sheet
#print(wb.sheetnames)

#Cria as planilhas das turmas a partir da planilha padrão
list = []
for d in wb["Turmas"].iter_rows(values_only=True,min_row=2):
    if (d[0]!= None):
        list.append(d[0])
        wb.copy_worksheet(wb['Padrao']).title = d[0]
        wb[d[0]]['k1']  = d[0]
        wb[d[0]]['J6'] = '=QUERY(Disciplinas!A2:J493;"select J where B = \'"&k1&"\' ")'
        #Validação
        for validation in wb['Padrao'].data_validations.dataValidation:
            wb[d[0]].add_data_validation(copy(validation))
        #Formatação condicional
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(font=red_text, fill=red_fill)
        rule = Rule(type="containsText", operator="containsText", text="A concluir", dxf=dxf)
        #rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
        rule.formula = ['=$M6="A concluir"']
        wb[d[0]].conditional_formatting.add('J6:M800', rule)
       
#Atualiza Formulas do relatório professor
columns = fruits = ["C", "D", "E","F","G","H"]
txt = "=IF(REGEXMATCH("+list[0]+"!{column}{row} ;$C$3); 1; 0)"
for item in list[1:]:
    txt += "+ IF(REGEXMATCH("+item+"!{column}{row} ;$C$3); 1; 0)"
for column in columns:
    for x in range(3, 19):
        wb['Relatório Professor'][column+str(x+3)] = txt.format(column = column, row=x)





wb.save('files/copera_processed.xlsx')